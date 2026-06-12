"""
Route: Laporan
---------------
Halaman laporan booking — hanya HRD, admin, dan direktur.
Termasuk fitur export ke Excel.
"""
from datetime import date, datetime
from io import BytesIO
from flask import Blueprint, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.models.room_booking import RoomBooking
from app.models.meeting_room import MeetingRoom
from app.models.division import Division
from app.services.report_service import ReportService
from app.routes.auth import admin_or_hrd_required

report_bp = Blueprint('report', __name__, url_prefix='/report')


@report_bp.route('/')
@admin_or_hrd_required
def index():
    """Halaman laporan dengan filter."""
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    division = request.args.get('division', '')
    room_id = request.args.get('room_id', type=int)

    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    summary = ReportService.get_summary(
        start_date=start_date, end_date=end_date,
        division=division if division else None, room_id=room_id)

    rooms = MeetingRoom.get_active_rooms()
    divisions = Division.get_active_choices()

    return render_template('report/index.html',
        summary=summary, rooms=rooms, divisions=divisions,
        start_date=start_date_str, end_date=end_date_str,
        division_filter=division, room_id_filter=room_id)


@report_bp.route('/export')
@admin_or_hrd_required
def export_excel():
    """Export laporan ke file Excel."""
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    division = request.args.get('division', '')
    room_id = request.args.get('room_id', type=int)
    status = request.args.get('status', '')

    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    bookings = ReportService.get_export_data(
        start_date=start_date, end_date=end_date,
        division=division if division else None, room_id=room_id,
        status=status if status else None)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Laporan Booking'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['No', 'Judul Meeting', 'Pengaju', 'Divisi', 'Ruangan',
                'Tanggal', 'Jam Mulai', 'Jam Selesai', 'Peserta',
                'Keperluan', 'Status', 'Approver', 'Tanggal Approve']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, booking in enumerate(bookings, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        ws.cell(row=row_idx, column=2, value=booking.title).border = thin_border
        ws.cell(row=row_idx, column=3, value=booking.user.nama).border = thin_border
        ws.cell(row=row_idx, column=4, value=booking.division).border = thin_border
        ws.cell(row=row_idx, column=5, value=booking.room.room_name).border = thin_border
        ws.cell(row=row_idx, column=6, value=str(booking.meeting_date)).border = thin_border
        ws.cell(row=row_idx, column=7, value=str(booking.start_time)[:5]).border = thin_border
        ws.cell(row=row_idx, column=8, value=str(booking.end_time)[:5]).border = thin_border
        ws.cell(row=row_idx, column=9, value=booking.participant_count).border = thin_border
        ws.cell(row=row_idx, column=10, value=booking.purpose).border = thin_border
        ws.cell(row=row_idx, column=11, value=booking.status).border = thin_border
        ws.cell(row=row_idx, column=12, value=booking.approver.nama if booking.approver else '').border = thin_border
        ws.cell(row=row_idx, column=13, value=str(booking.approved_at)[:19] if booking.approved_at else '').border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'laporan_booking_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename)